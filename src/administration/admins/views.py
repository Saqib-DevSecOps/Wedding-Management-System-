import openpyxl
from django.contrib.auth.decorators import login_required, user_passes_test
from django.core.paginator import Paginator
from django.shortcuts import get_object_or_404, redirect, render
from django.contrib import messages
from django.urls import reverse_lazy
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_GET
from django.views.generic import (
    TemplateView, ListView, DetailView, UpdateView, CreateView
)
from openpyxl.utils import get_column_letter

from src.administration.admins.filters import ProviderFilter, TableFilter
from src.administration.admins.forms import GuestGroupMetaForm, ProviderMetaForm, \
    TableForm, EventTimeLineMetaForm
from src.administration.admins.models import GuestGroup, Guest, Provider, InvitationLetter, Table, GuestTable, \
    EventTimeLine
from django.views import View
import json
from django.http import JsonResponse
import os
from django.conf import settings
from django.http import HttpResponse, Http404
from django.core.serializers.json import DjangoJSONEncoder

admin_decorators = [login_required, user_passes_test(lambda u: u.is_superuser)]


@method_decorator(login_required, name='dispatch')
class DashboardView(TemplateView):
    template_name = 'admins/dashboard.html'

    def get_context_data(self, **kwargs):
        context = super(DashboardView, self).get_context_data(**kwargs)

        total_group = GuestGroup.objects.filter(user=self.request.user).count()
        provider = Provider.objects.filter(user=self.request.user)

        total_payment = 0
        paid_payment = 0

        for objects in provider:
            if objects.total_cost is not None:
                total_payment += objects.total_cost

            if objects.paid is not None:
                paid_payment += objects.paid

        remaining_payment = total_payment - paid_payment

        context['recent_events'] = EventTimeLine.objects.filter(user=self.request.user).order_by('-date')[:3]
        context['recent_seats'] = Table.objects.filter(user=self.request.user)[:6]
        context['recent_groups'] = GuestGroup.objects.filter(user=self.request.user)[:6]
        context['recent_providers'] = Provider.objects.filter(user=self.request.user)[:6]

        context['total_events'] = EventTimeLine.objects.filter(user= self.request.user).count()
        context['total_payment'] = total_payment
        context['paid_payment'] = paid_payment
        context['remaining_payment'] = remaining_payment
        context['total_provider'] = provider.count()

        total_invitation_letters = 0
        invitation_letters = InvitationLetter.objects.filter(group__user=self.request.user)
        for invitation in invitation_letters:
            total_invitation_letters += invitation.total_invitation

        context['total_group'] = total_group
        context['total_invitations'] = total_invitation_letters
        context['total_guests'] = Guest.objects.filter(group__user=self.request.user).count()

        table = Table.objects.filter(user=self.request.user)
        rounded = table.filter(table_type='1').count()
        rectangle = table.filter(table_type='2').count()

        context['total_table'] = table.count()
        context['rounded'] = rounded
        context['rectangle'] = rectangle

        return context


@method_decorator(login_required, name='dispatch')
class GuestGroupListView(CreateView, ListView):
    model = GuestGroup
    form_class = GuestGroupMetaForm
    template_name = 'admins/guest_group_list.html'
    success_url = reverse_lazy("admins:guest-group-list")

    def get_queryset(self):
        search = self.request.GET.get('guest_group_search')
        if search:
            return self.model.objects.filter(user=self.request.user, group_name__icontains=search)
        return self.model.objects.filter(user=self.request.user)

    def get_context_data(self,*args, **kwargs):
        context = super(GuestGroupListView, self).get_context_data(**kwargs)
        context['invitation'] = InvitationLetter.objects.filter(group__user=self.request.user)
        return context

    def form_valid(self, form):
        guest_names = self.request.POST.getlist('guest_names[]')
        guest_group = form.save(commit=False)  # Get the saved GuestGroup instance
        guest_group.user = self.request.user
        guest_group.save()
        invitation_letter, created = InvitationLetter.objects.get_or_create(group=guest_group)
        invitation_letter.save()
        # Save guest names to the guest group
        for name in guest_names:
            Guest.objects.create(guest_name=name, group=guest_group)
        messages.success(self.request, 'Guest Group Created Successfully')
        return super().form_valid(form)




@csrf_exempt
def save_guest_group(request):
    if request.method == 'POST':
        group_name = request.POST.get('group_name')
        guest_names = request.POST.getlist('guest_names[]')
        guest_group = GuestGroup.objects.create(group_name=group_name, user=request.user)
        for name in guest_names:
            guest_group.guest_set.create(guest_name=name)
        InvitationLetter.objects.create(group=guest_group)

        return JsonResponse({'message': 'Guest group created successfully', 'group_id': guest_group.id})
    else:
        return JsonResponse({'error': 'Invalid request method. Only POST requests are allowed.'}, status=405)


@method_decorator(login_required, name='dispatch')
class GuestGroupDetailView(DetailView):
    model = GuestGroup
    template_name = 'admins/guest_group_detail.html'

    def get_object(self, queryset=None):
        return get_object_or_404(GuestGroup, pk=self.kwargs['pk'], user=self.request.user)


@login_required
def update_guest_group(request):
    if request.method == "GET":
        group_id = request.GET.get('group_id')
        guest_group = get_object_or_404(GuestGroup, id=group_id, user=request.user)

        # Prepare the data to be sent in the response
        response_data = {
            'success': True,
            'group_name': guest_group.group_name,
            'guest_names': list(guest_group.guest_set.values_list('guest_name', flat=True)),
            'guest_ids': list(guest_group.guest_set.values_list('id', flat=True)),
        }

        return JsonResponse(response_data)

    if request.method == "POST":
        group_id = request.POST.get('group_id')
        group_name = request.POST.get('group_name')
        guest_names = request.POST.getlist('guest_names[]')
        # Fetch the guest group object
        guest_group = get_object_or_404(GuestGroup, id=group_id, user=request.user)
        # Update the guest group data
        guest_group.group_name = group_name
        guest_group.save()

        # Update the guest names
        guest_group.guest_set.all().delete()  # Remove existing guest names
        for name in guest_names:
            guest_group.guest_set.create(guest_name=name)
        messages.success(request, "Group Updated Successfully")
        return JsonResponse({'success': True})


@method_decorator(login_required, name='dispatch')
class GuestGroupUpdateView(UpdateView):
    model = GuestGroup
    form_class = GuestGroupMetaForm
    template_name = 'admins/guest_group_update.html'
    success_url = reverse_lazy('admins:guest-group-list')

    def form_valid(self, form):
        guest_group = form.save()
        # Handle saving the updated guest list if necessary
        return JsonResponse({'success': True})


class InvitationUpdateView(View):

    def get(self, request, *args, pk):
        invitation = get_object_or_404(InvitationLetter, id=pk)
        data = {
            'total_invitation': invitation.total_invitation,
        }
        return JsonResponse({'invitation': data})

    def post(self, request, pk):
        invitation = get_object_or_404(InvitationLetter, id=pk)
        total = request.POST.get('total_invitation')
        if total:
            invitation.total_invitation = total
            invitation.save()
            messages.success(request, "Invitation Letter Successfully updated")
            return JsonResponse({'success': True, 'provider_id': invitation.id})

        errors = {}
        for field, error_messages in form.errors.items():
            errors[field] = [str(message) for message in error_messages]

        return JsonResponse({'errors': errors}, status=400)


@method_decorator(login_required, name='dispatch')
class GuestGroupDeleteView(View):
    def get(self, request, pk):
        guest = get_object_or_404(GuestGroup, pk=pk)
        guest.delete()
        messages.success(self.request, "Guest Group Successfully Deleted")
        return redirect('admins:guest-group-list')


@require_GET
def get_guests(request):
    group_id = request.GET.get('group_id')
    group = get_object_or_404(GuestGroup, id=group_id)
    guests = Guest.objects.filter(group=group)
    guest_list = [{'name': guest.guest_name} for guest in guests]
    return JsonResponse(guest_list, safe=False)


@method_decorator(login_required, name='dispatch')
class ProviderListCreateView(ListView):
    model = Provider
    template_name = 'admins/provider_list.html'
    success_url = reverse_lazy("admins:guest-group-list")

    def get_queryset(self):
        queryset = super().get_queryset()
        return queryset.filter(user=self.request.user)

    def get_context_data(self, **kwargs):
        context = super(ProviderListCreateView, self).get_context_data(**kwargs)
        context['form'] = ProviderMetaForm
        filter_object = ProviderFilter(self.request.GET, queryset=self.get_queryset())
        object_list = filter_object.qs
        paginator = Paginator(object_list, 20)
        page_number = self.request.GET.get('page')
        page_object = paginator.get_page(page_number)
        context['object_list'] = page_object
        context['filter_form'] = filter_object.form
        return context


@method_decorator(login_required, name='dispatch')
class ProviderDetailView(DetailView):
    model = Provider
    template_name = 'admins/provider_detail.html'

    def get_object(self, queryset=None):
        return get_object_or_404(Provider, id=self.kwargs['pk'], user=self.request.user)


@method_decorator(login_required, name='dispatch')
class ProviderCreateView(View):
    def post(self, request):
        form = ProviderMetaForm(request.POST, request.FILES)
        if form.is_valid():
            provider = form.save(commit=False)
            provider.user = request.user
            provider.save()
            messages.success(request, "Successfully Created")
            return JsonResponse({'success': True, 'provider_id': provider.id})

        errors = {}
        for field, error_messages in form.errors.items():
            errors[field] = [str(message) for message in error_messages]

        return JsonResponse({'errors': errors}, status=400)


@method_decorator(login_required, name='dispatch')
class ProviderUpdateView(View):

    def get(self, request, pk):
        provider = get_object_or_404(Provider, id=pk, user=self.request.user)

        provider_data = {
            'provider_name': provider.provider_name,
            'service': provider.service,
            'email': provider.email,
            'link': provider.link,
            'phone_number': provider.phone_number,
            'total_cost': provider.total_cost,
            'paid': provider.paid,
            'comment': provider.comment,
        }

        return JsonResponse({'provider': provider_data}, encoder=DjangoJSONEncoder)

    def post(self, request, pk):
        provider = get_object_or_404(Provider, id=pk, user=self.request.user)
        form = ProviderMetaForm(request.POST, request.FILES, instance=provider)
        if form.is_valid():
            provider = form.save()
            messages.success(request, "Successfully updated")
            return JsonResponse({'success': True, 'provider_id': provider.id})

        errors = {}
        for field, error_messages in form.errors.items():
            ("invalid")

            errors[field] = [str(message) for message in error_messages]

        return JsonResponse({'errors': errors}, status=400)


@method_decorator(login_required, name='dispatch')
class ProviderDeleteView(View):
    def get(self, request, pk):
        provider = get_object_or_404(Provider, id=pk)
        provider.delete()
        messages.success(self.request, "Provider Successfully Deleted")
        return redirect('admins:provider-list')


@login_required
def update_row_order(request):
    if request.method == 'POST':
        body = json.loads(request.body)
        element_id = body.get('card_id_list', [])
        for index in range(len(element_id)):
            group = GuestGroup.objects.get(pk=int(element_id[index]))
            group.sequence = index
            group.save()
        return JsonResponse({'success': True})

    return JsonResponse({'error': 'Invalid request method'})


@login_required
def update_invitation_order(request):
    if request.method == 'POST':
        body = json.loads(request.body)
        element_id = body.get('row_id_list', [])
        for index in range(len(element_id)):
            invitation = InvitationLetter.objects.get(pk=int(element_id[index]))
            invitation.sequence = index
            invitation.save()
        return JsonResponse({'success': True})

    return JsonResponse({'error': 'Invalid request method'})


@method_decorator(login_required, name='dispatch')
class SeatPlannerListView(ListView):
    model = Table
    template_name = 'admins/seat_planner_list.html'

    def get_queryset(self):
        return super().get_queryset().filter(user=self.request.user)

    def get_context_data(self, **kwargs):
        context = super(SeatPlannerListView, self).get_context_data(**kwargs)
        filter_object = TableFilter(self.request.GET, queryset=self.get_queryset())

        object_list = filter_object.qs
        paginator = Paginator(object_list, 20)
        page_number = self.request.GET.get('page')
        page_object = paginator.get_page(page_number)
        context['object_list'] = page_object

        context['filter_form'] = filter_object.form
        return context


@method_decorator(login_required, name='dispatch')
class SeatPlannerCreateView(TemplateView):
    template_name = 'admins/create_seat_planner.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['guests'] = Guest.objects.filter(group__user=self.request.user)
        context['user'] = self.request.user.id
        context['form'] = TableForm
        return context


@method_decorator(login_required, name='dispatch')
class CreateSeatPlannerViewApi(View):
    def post(self, request, *args, **kwargs):
        try:
            container_records = json.loads(request.body)
            for i in range(len(container_records)):
                seat_count = int(container_records[i]['table_size'])
                table = Table(user=self.request.user, table_name=container_records[i]['table_name'],
                              seat_count=seat_count,
                              table_type=container_records[i]['table_type'],
                              )
                table.seats_left = int(seat_count - len(container_records[i]['guests']))
                table.save()
                for j in range(len(container_records[i]['guests'])):
                    guest_table = GuestTable(table=table, guest_id=container_records[i]['guests'][j])
                    guest_table.save()
            messages.success(request, "Table Designed Successfully")
            return JsonResponse({'message': 'Data received and processed successfully.'})
        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON data.'}, status=400)

    def http_method_not_allowed(self, request, *args, **kwargs):
        return JsonResponse({'error': 'Invalid request method.'}, status=405)


@method_decorator(login_required, name='dispatch')
class SeatPlannerDetail(DetailView):
    model = Table
    template_name = 'admins/seat_planner_detail.html'


@method_decorator(login_required, name='dispatch')
class SeatPlannerDelete(View):
    def get(self, request, pk):
        table = get_object_or_404(Table, id=pk)
        table.delete()
        messages.success(request, 'Table Deleted Successfully')
        return redirect('admins:seat-planner-list')


@method_decorator(login_required, name='dispatch')
class UpdateSeatPlanner(DetailView):
    model = Table
    template_name = 'admins/update_seat_planner.html'

    def get_object(self, queryset=None):
        return get_object_or_404(Table, id=self.kwargs['pk'], user=self.request.user)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form'] = TableForm
        context['guests'] = Guest.objects.filter(group__user=self.request.user)
        return context


@method_decorator(login_required, name='dispatch')
class DownloadAttachmentView(View):
    def get(self, request, provider_id):
        try:
            provider = Provider.objects.get(id=provider_id)
        except Provider.DoesNotExist:
            raise Http404

        if not provider.attachment:
            raise Http404

        attachment_path = os.path.join(settings.MEDIA_ROOT, str(provider.attachment))
        if os.path.exists(attachment_path):
            with open(attachment_path, 'rb') as attachment_file:
                response = HttpResponse(
                    attachment_file.read(),
                    content_type='application/octet-stream',
                )
                response['Content-Disposition'] = 'attachment; filename="{}"'.format(provider.attachment.name)
                return response

        raise Http404


@method_decorator(login_required, name='dispatch')
class ExportGroupsToExcel(View):
    def get(self, request, *args, **kwargs):
        user = request.user
        groups = GuestGroup.objects.filter(user=user)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="user_groups_export.xlsx"'

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Groups Data'

        # Writing headers
        columns = ['Group Name', 'Guest Name', 'Total Invitations']
        for col_num, column_title in enumerate(columns, 1):
            col_letter = get_column_letter(col_num)
            ws[f'{col_letter}1'] = column_title

        # Writing data for each group
        row_num = 2
        for group in groups:
            ws[f'A{row_num}'] = group.group_name
            ws[f'C{row_num}'] = group.invitation_set.total_invitation

            guests = group.guest_set.all()
            for guest in guests:
                ws[f'B{row_num}'] = guest.guest_name
                row_num += 1

        for col_num in range(1, len(columns) + 1):
            col_letter = get_column_letter(col_num)
            ws.column_dimensions[col_letter].auto_size = True

        wb.save(response)
        return response


@method_decorator(login_required, name='dispatch')
class EventTimelineView(ListView):
    model = EventTimeLine
    template_name = 'admins/event.html'

    def get_queryset(self):
        queryset = super().get_queryset()
        return queryset.filter(user=self.request.user)

    def get_context_data(self, **kwargs):
        context = super(EventTimelineView, self).get_context_data(**kwargs)
        context['form'] = EventTimeLineMetaForm
        object_list = self.get_queryset()
        paginator = Paginator(object_list, 20)
        page_number = self.request.GET.get('page')
        page_object = paginator.get_page(page_number)
        context['object_list'] = page_object
        return context


@method_decorator(login_required, name='dispatch')
class EventTimeLineCreateView(View):
    def post(self, request):
        form = EventTimeLineMetaForm(request.POST)
        if form.is_valid():
            event = form.save(commit=False)
            event.user = request.user
            event.save()
            messages.success(request, "Event Successfully Created")
            return JsonResponse({'success': True, 'provider_id': event.id})

        errors = {}
        for field, error_messages in form.errors.items():
            errors[field] = [str(message) for message in error_messages]

        return JsonResponse({'errors': errors}, status=400)


@method_decorator(login_required, name='dispatch')
class EventTimeLineDeleteView(View):
    def get(self, request, pk):
        event = get_object_or_404(EventTimeLine, id=pk)
        event.delete()
        messages.success(self.request, "Event Successfully Deleted")
        return redirect('admins:event-timeline')


@method_decorator(login_required, name='dispatch')
class EventTimeLineUpdateView(View):

    def get(self, request, pk):
        event = get_object_or_404(EventTimeLine, id=pk, user=self.request.user)
        provider_data = {
            'title': event.title,
            'date': event.date.strftime('%Y-%m-%dT%H:%M'),  # Format the date correctly
            'description': event.description,
        }

        return JsonResponse({'event': provider_data}, encoder=DjangoJSONEncoder)

    def post(self, request, pk):
        event = get_object_or_404(EventTimeLine, id=pk, user=self.request.user)
        form = EventTimeLineMetaForm(request.POST, instance=event)
        if form.is_valid():
            provider = form.save()
            messages.success(request, "Event Successfully updated")
            return JsonResponse({'success': True, 'event_id': event.id})

        errors = {}
        for field, error_messages in form.errors.items():
            errors[field] = [str(message) for message in error_messages]

        return JsonResponse({'errors': errors}, status=400)


class Test(TemplateView):
    template_name = 'admins/test.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['guests'] = Guest.objects.filter(group__user=self.request.user)
        context['form'] = TableForm
        return context


